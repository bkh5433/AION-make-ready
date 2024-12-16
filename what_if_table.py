from typing import Dict, Tuple, List, Optional, Union, Any
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
import re
from logger_config import LogConfig

# Initialize logging

log_config = LogConfig()
logger = log_config.get_logger('what_if_table')


class FormulaEvaluator:
    """Handles evaluation of Excel formulas"""

    @staticmethod
    def _get_range_values(sheet, range_str: str) -> list:
        """Get values from a cell range (e.g., 'A1:A3')."""
        try:
            start_cell, end_cell = range_str.split(':')
            # Extract column letters and row numbers
            start_col = ''.join(c for c in start_cell if c.isalpha())
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            end_col = ''.join(c for c in end_cell if c.isalpha())
            end_row = int(''.join(c for c in end_cell if c.isdigit()))

            values = []
            for row in range(start_row, end_row + 1):
                # For now, assuming we're only dealing with single column ranges
                cell_value = sheet[f"{start_col}{row}"].value
                if cell_value is not None:
                    values.append(float(cell_value))

            return values
        except Exception as e:
            logger.error(f"Error processing range {range_str}: {str(e)}")
            return []

    @staticmethod
    def evaluate_formula(formula: str, sheet) -> float:
        """Evaluate an Excel formula."""
        try:
            # Remove the leading '=' if present
            formula = formula.lstrip('=')

            # Handle Excel AVERAGE function
            if 'AVERAGE' in formula:
                # Extract the range from AVERAGE(range)
                range_start = formula.find('(') + 1
                range_end = formula.find(')')
                if range_start > 0 and range_end > range_start:
                    range_str = formula[range_start:range_end]
                    values = FormulaEvaluator._get_range_values(sheet, range_str)
                    if values:
                        return sum(values) / len(values)
                    return 0

            # Handle other cell references
            cell_refs = re.findall(r'[A-Z]+[0-9]+', formula)

            # Replace cell references with their values
            for cell_ref in cell_refs:
                cell_value = sheet[cell_ref].value
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_value = FormulaEvaluator.evaluate_formula(cell_value, sheet)
                if cell_value is None:
                    cell_value = 0
                formula = formula.replace(cell_ref, str(float(cell_value)))

            # Evaluate the expression
            result = eval(formula)
            return float(result)

        except Exception as e:
            logger.error(f"Error evaluating formula {formula}: {e}")
            return 0


class WhatIfTableGenerator:
    """Handles generation and formatting of the what-if table"""

    def __init__(self, worksheet: Worksheet):
        self.sheet = worksheet
        self.start_row = 12  # Starting row for the what-if table
        self.header_row = 10
        self.daily_col = 'F'
        self.monthly_col = 'G'
        self.label_col = 'H'
        self.days_per_month = 21.7  # Standard working days per month
        self.formula_evaluator = FormulaEvaluator()

    def _get_border(self, position='middle') -> Border:
        """
        Get border style based on position in table.
        position can be: 'top', 'bottom', 'middle'
        """
        thin = Side(style='thin')
        thick = Side(style='medium')  # Using medium for outer borders

        if position == 'top':
            return Border(top=thick, bottom=thin, left=thick, right=thick)
        elif position == 'bottom':
            return Border(top=thin, bottom=thick, left=thick, right=thick)
        else:  # middle
            return Border(top=thin, bottom=thin, left=thick, right=thick)

    def _format_headers(self):
        """Format the header row of the table"""
        # Header text
        headers = {
            self.daily_col: 'Daily work orders completed',
            self.monthly_col: 'Work orders completed per-month',
            # self.label_col: 'Current monthly output'
        }

        # Style headers
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # Light blue
        header_font = Font(name='Aptos Narrow Bold', bold=True)

        for col, text in headers.items():
            cell = self.sheet[f'{col}{self.header_row}']
            cell.value = text
            cell.fill = header_fill
            cell.font = header_font
            cell.border = self._get_border('top')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def _create_cell_style(self,
                           bg_color: str,
                           font_color: str = "000000",
                           bold: bool = False,
                           position='middle') -> Tuple[PatternFill, Font, Border]:
        """Create cell styling components."""
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        font = Font(name='Aptos Narrow Body', color=font_color, bold=bold)
        border = self._get_border(position)
        return fill, font, border

    def _highlight_row(self,
                       row: int,
                       label: str,
                       bg_color: str,
                       font_color: str = "000000",
                       bold: bool = False,
                       position='middle') -> None:
        """Apply highlighting to a specific row."""
        fill, font, border = self._create_cell_style(bg_color, font_color, bold, position)

        for col in [self.daily_col, self.monthly_col]:
            cell = self.sheet[f'{col}{row}']
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

        label_cell = self.sheet[f'{self.label_col}{row}']
        label_cell.value = label
        label_cell.fill = fill
        label_cell.font = font
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left')

    def _calculate_monthly_rate(self, daily_rate: float) -> float:
        """Calculate monthly rate from daily rate."""
        return round(daily_rate * self.days_per_month, 1)

    def calculate_metrics(self, data: Dict) -> Dict[str, float]:
        """Calculate metrics from sheet formulas and data."""
        logger.info("Calculating metrics from sheet data")

        # Populate necessary cells
        self.sheet['B6'] = 21
        self.sheet['M9'] = data.get('ActualOpenWorkOrders_Current', 0)

        # Evaluate break-even formula
        break_even_formula = self.sheet['B24'].value
        break_even_value = (
            self.formula_evaluator.evaluate_formula(break_even_formula, self.sheet)
            if isinstance(break_even_formula, str) and break_even_formula.startswith('=')
            else float(break_even_formula or 0)
        )

        # Evaluate current output formula
        current_output_formula = self.sheet['B4'].value
        current_output_value = (
            self.formula_evaluator.evaluate_formula(current_output_formula, self.sheet)
            if isinstance(current_output_formula, str) and current_output_formula.startswith('=')
            else float(current_output_formula or 0)
        )

        logger.info(f"Calculated metrics - Break Even: {break_even_value}, Current Output: {current_output_value}")

        return {
            'break_even_value': break_even_value,
            'current_output_value': current_output_value
        }

    def generate_table(self,
                       current_output: float,
                       break_even_target: float,
                       increment: float = 1.0) -> None:
        """Generate the what-if table with current output and break-even target."""
        logger.info(f"Generating what-if table - Current Output: {current_output}, Break Even: {break_even_target}")

        try:
            # Format headers first
            self._format_headers()
            # Round values for consistent comparison
            current_output = round(current_output, 1)
            break_even_target = round(break_even_target, 1)

            current_row = self.start_row
            found_current = False
            found_break_even = False

            # Generate table rows starting from 1 to 52
            for i in range(1, 53):  # Changed to start from 1 and go to 52
                daily_rate = i  # Remove decimal places for regular rows
                monthly_rate = round(daily_rate * self.days_per_month)  # Round to whole number

                # Determine if this is the last row
                is_last_row = i == 52
                position = 'bottom' if is_last_row else 'middle'

                # Format cells
                daily_cell = self.sheet[f'{self.daily_col}{current_row}']
                monthly_cell = self.sheet[f'{self.monthly_col}{current_row}']

                # Set values
                daily_cell.value = daily_rate
                monthly_cell.value = monthly_rate

                # Set number format
                daily_cell.number_format = '0'  # Display as whole number
                monthly_cell.number_format = '0'  # Display as whole number

                # Basic cell styling
                for cell in [daily_cell, monthly_cell]:
                    cell.font = Font(name='Aptos Narrow Body')
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = self._get_border(position)

                # Check if we need to insert current output before this row
                if not found_current and daily_rate > current_output:
                    # Insert current output row
                    self._highlight_row(
                        current_row,
                        "<-------- Current output (AVG)",
                        "FFA500",  # Orange
                        "000000",  # Black text
                        True,
                        position
                    )
                    daily_cell.value = current_output
                    monthly_cell.value = self._calculate_monthly_rate(current_output)
                    daily_cell.number_format = '0.0'
                    monthly_cell.number_format = '0.0'
                    found_current = True
                    current_row += 1

                    # Re-create the regular row we were processing
                    daily_cell = self.sheet[f'{self.daily_col}{current_row}']
                    monthly_cell = self.sheet[f'{self.monthly_col}{current_row}']
                    daily_cell.value = daily_rate
                    monthly_cell.value = monthly_rate
                    daily_cell.number_format = '0'
                    monthly_cell.number_format = '0'
                    for cell in [daily_cell, monthly_cell]:
                        cell.font = Font(name='Aptos Narrow Body')
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = self._get_border(position)

                # Check if we need to insert break-even target
                if not found_break_even and daily_rate > break_even_target:
                    self._highlight_row(
                        current_row,
                        "<-------- Break even",
                        "0000FF",  # Blue
                        "FFFFFF",  # White text
                        True,
                        position
                    )
                    daily_cell.value = break_even_target
                    monthly_cell.value = self._calculate_monthly_rate(break_even_target)
                    daily_cell.number_format = '0.0'
                    monthly_cell.number_format = '0.0'
                    found_break_even = True
                    current_row += 1

                    # Re-create the regular row we were processing if not the last row
                    if not is_last_row:
                        daily_cell = self.sheet[f'{self.daily_col}{current_row}']
                        monthly_cell = self.sheet[f'{self.monthly_col}{current_row}']
                        daily_cell.value = daily_rate
                        monthly_cell.value = monthly_rate
                        daily_cell.number_format = '0'
                        monthly_cell.number_format = '0'
                        for cell in [daily_cell, monthly_cell]:
                            cell.font = Font(name='Aptos Narrow Body')
                            cell.alignment = Alignment(horizontal='right')
                            cell.border = self._get_border(position)

                current_row += 1

            # If we haven't found current output or break-even by the end, add them at the last row
            if not found_current or not found_break_even:
                if not found_current:
                    self._highlight_row(
                        current_row - 1,
                        "<-------- Current output (AVG)",
                        "FFA500",
                        "000000",
                        True,
                        'bottom'
                    )
                if not found_break_even:
                    self._highlight_row(
                        current_row - 1,
                        "<-------- Break even",
                        "0000FF",
                        "FFFFFF",
                        True,
                        'bottom'
                    )

            logger.info("What-if table generated successfully")

        except Exception as e:
            logger.error(f"Error generating what-if table: {str(e)}", exc_info=True)
            raise


def update_what_if_table(sheet: Worksheet,
                         data: Dict[str, Any],
                         ) -> None:
    """
    Update the what-if table in the worksheet.

    Args:
        sheet: Worksheet to update
        data: Property data dictionary
        open_actual: Actual open work orders
    """
    try:
        generator = WhatIfTableGenerator(sheet)

        # Calculate metrics first
        metrics = generator.calculate_metrics(data)

        # Generate the table
        generator.generate_table(
            current_output=metrics['current_output_value'],
            break_even_target=metrics['break_even_value']
        )

        logger.info("What-if table updated successfully")

    except Exception as e:
        logger.error(f"Failed to update what-if table: {str(e)}", exc_info=True)
        raise


if __name__ == "__main__":
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    from pathlib import Path
    import os
    import tempfile
    import shutil

    test_logger = log_config.get_logger('what_if_table_tests')


    def run_tests():
        test_logger.info("\n=== Running What-If Table Tests ===\n")

        # Create a temporary directory for test files
        temp_dir = tempfile.mkdtemp()
        try:
            # Test FormulaEvaluator
            def test_formula_evaluator():
                test_logger.info("Testing FormulaEvaluator...")

                # Create a test workbook
                wb = openpyxl.Workbook()
                sheet = wb.active

                # Setup test data
                sheet['A1'] = 10.0
                sheet['A2'] = 20.0
                sheet['A3'] = 30.0

                evaluator = FormulaEvaluator()

                # Test simple formula
                test_logger.info("Testing simple addition...")
                result1 = evaluator.evaluate_formula('=A1+A2', sheet)
                assert result1 == 30, f"Simple formula test failed: expected 30, got {result1}"
                test_logger.info("✓ Simple addition test passed")

                # Test AVERAGE function
                test_logger.info("Testing AVERAGE function...")
                result2 = evaluator.evaluate_formula('=AVERAGE(A1:A3)', sheet)
                assert abs(result2 - 20.0) < 0.01, f"AVERAGE function test failed: expected 20, got {result2}"
                test_logger.info("✓ AVERAGE function test passed")

                # Test multiplication
                test_logger.info("Testing multiplication...")
                result3 = evaluator.evaluate_formula('=A1*2', sheet)
                assert result3 == 20, f"Multiplication test failed: expected 20, got {result3}"
                test_logger.info("✓ Multiplication test passed")

                test_logger.info("✓ All FormulaEvaluator tests passed")

            def test_what_if_table_generator():
                test_logger.info("\nTesting WhatIfTableGenerator...")

                # Create a test workbook with template-like structure
                wb = openpyxl.Workbook()
                sheet = wb.active

                # Setup test data
                sheet['B4'] = '=M9/21'  # Current output formula
                sheet['B24'] = 25  # Break-even target
                sheet['M9'] = 420  # Open work orders

                generator = WhatIfTableGenerator(sheet)

                # Test metrics calculation
                test_logger.info("Testing metrics calculation...")
                metrics = generator.calculate_metrics({}, 420)
                assert 'break_even_value' in metrics, "Break-even value missing from metrics"
                assert 'current_output_value' in metrics, "Current output value missing from metrics"
                test_logger.info("✓ Metrics calculation test passed")

                # Test table generation
                test_logger.info("Testing table generation...")
                generator.generate_table(
                    current_output=20,
                    break_even_target=25
                )

                # Verify table contents
                # Check first row (current output)
                daily_value = sheet['F11'].value
                monthly_value = sheet['G11'].value
                assert daily_value == 20, f"Initial daily value incorrect: expected 20, got {daily_value}"
                assert round(monthly_value, 1) == round(20 * 21.7, 1), f"Initial monthly value incorrect"
                test_logger.info("✓ Table generation test passed")

                # Save test output for inspection
                test_output_path = Path(temp_dir) / "test_output.xlsx"
                wb.save(test_output_path)
                test_logger.info(f"\nTest output saved to: {test_output_path}")

            def test_complete_workflow():
                test_logger.info("\nTesting complete workflow...")

                # Create a test workbook
                wb = openpyxl.Workbook()
                sheet = wb.active

                # Setup test data
                test_data = {
                    'ActualOpenWorkOrders_Current': 450,
                    'CancelledWorkOrder_Current': 30,
                }

                # Test the main update function
                try:
                    update_what_if_table(
                        sheet=sheet,
                        data=test_data,
                        open_actual=420
                    )
                    test_logger.info("✓ Complete workflow test passed")
                except Exception as e:
                    test_logger.info(f"✗ Complete workflow test failed: {str(e)}")
                    raise

            # Run all tests
            try:
                test_formula_evaluator()
                test_what_if_table_generator()
                test_complete_workflow()
                test_logger.info("\n✓ All tests passed successfully!")
            except AssertionError as e:
                test_logger.info(f"\n✗ Test failed: {str(e)}")
            except Exception as e:
                test_logger.info(f"\n✗ Unexpected error during testing: {str(e)}")

        finally:
            # Cleanup temporary directory
            shutil.rmtree(temp_dir)


    # Run the tests
    run_tests()
