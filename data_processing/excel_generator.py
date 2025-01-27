import shutil
import time
from pathlib import Path
from typing import Dict, List, Union, Any, Optional
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles.styleable import copy
from openpyxl.styles import Font, Alignment
from pydantic import BaseModel, Field, ConfigDict, computed_field
from what_if_table import update_what_if_table

from logger_config import LogConfig, log_exceptions
from utils.path_resolver import PathResolver
from models.Property import Property

# Initialize logging
log_config = LogConfig()
logger = log_config.get_logger("excel_generator")


class SpreadsheetTemplate(BaseModel):
    """Model for spreadsheet template configuration"""
    template_name: str = Field(..., description="Path to template file")
    output_path: Path = Field(..., description="Path where output file will be saved")
    sheet_name: str = Field(default="Sheet1", description="Name of the worksheet to use")

    @property
    def template_path(self) -> Path:
        """Resolve the template path dynamically"""
        try:
            return PathResolver.resolve_template_path(self.template_name)
        except FileNotFoundError as e:
            logger.error(f"Failed to resolve template path: {e}")
            raise

    model_config = ConfigDict(validate_assignment=True)


class WorkOrderMetrics(BaseModel):
    """Model for work order metrics with calculations"""
    # Work order tracking fields
    open_work_orders: int = Field(ge=0)
    actual_open_work_orders: int = Field(ge=0)
    new_work_orders: int = Field(ge=0, default=0)
    completed_work_orders: int = Field(ge=0)
    cancelled_work_orders: int = Field(ge=0)
    pending_work_orders: int = Field(ge=0)
    percentage_completed: float = Field(ge=0, le=100)
    average_days_to_complete: float = Field(ge=0, default=0.0)

    # Period dates
    period_start_date: Optional[datetime] = None
    period_end_date: Optional[datetime] = None

    # Configuration
    days_per_month: int = Field(ge=0, le=31, default=21)

    model_config = ConfigDict(
        validate_assignment=True,
        extra='allow',
        arbitrary_types_allowed=True
    )

    @computed_field
    def daily_rate(self) -> float:
        """Calculate daily completion rate"""
        try:
            return round(self.completed_work_orders / self.days_per_month, 1)
        except ZeroDivisionError:
            return 0.0

    @computed_field
    def monthly_rate(self) -> float:
        """Calculate monthly completion rate"""
        return round(self.daily_rate * self.days_per_month, 1)

    @computed_field
    def break_even_target(self) -> float:
        """Calculate break-even target with 10% buffer"""
        try:
            target_rate = max(self.actual_open_work_orders, self.completed_work_orders) / self.days_per_month
            return round(target_rate * 1.1, 1)
        except ZeroDivisionError:
            return 0.0

    @computed_field
    def current_output(self) -> float:
        """Current output matches daily rate"""
        return self.daily_rate

    def get_metrics_for_table(self) -> dict:
        """Get metrics formatted for what-if table calculations"""
        return {
            'daily_rate': self.daily_rate,
            'monthly_rate': self.monthly_rate,
            'break_even_target': self.break_even_target,
            'current_output': self.current_output,
            'days_per_month': self.days_per_month
        }

    def get_performance_metrics(self) -> dict:
        """Get comprehensive performance metrics"""
        return {
            **self.get_metrics_for_table(),
            'open_work_orders': self.open_work_orders,
            'actual_open_work_orders': self.actual_open_work_orders,
            'new_work_orders': self.new_work_orders,
            'completed_work_orders': self.completed_work_orders,
            'cancelled_work_orders': self.cancelled_work_orders,
            'pending_work_orders': self.pending_work_orders,
            'percentage_completed': self.percentage_completed,
            'average_days_to_complete': self.average_days_to_complete
        }


class ExcelGeneratorService:
    """Service for handling Excel generation with improved error handling and logging"""

    def __init__(self, template: SpreadsheetTemplate):
        """Initialize the Excel generator service"""
        logger.info(f"Initializing ExcelGeneratorService with template: {template.template_path}")
        self.template = template
        self.workbook = None
        self.sheet = None
        self._font = Font(name='Aptos Narrow Bold')

    @log_exceptions(logger)
    def initialize_workbook(self) -> None:
        """Initialize workbook from template with comprehensive error handling"""
        logger.info(f"Loading template from: {self.template.template_path}")
        try:
            self.workbook = openpyxl.load_workbook(self.template.template_path)
            if self.template.sheet_name == "Sheet1":
                self.sheet = self.workbook.active
            else:
                if self.template.sheet_name not in self.workbook.sheetnames:
                    raise ValueError(f"Sheet '{self.template.sheet_name}' not found in workbook")
                self.sheet = self.workbook[self.template.sheet_name]

            logger.info(f"Successfully loaded workbook with {len(self.workbook.sheetnames)} sheets")
        except Exception as e:
            logger.error(f"Failed to initialize workbook", exc_info=True)
            raise ValueError(f"Failed to initialize workbook: {str(e)}") from e

    @log_exceptions(logger)
    def update_cell(self, cell_ref: str, value: Any, wrap_text: bool = False) -> None:
        """Update a cell with proper formatting and error handling"""
        if not self.sheet:
            logger.error("Attempted to update cell without initialized sheet")
            raise ValueError("Sheet not initialized")

        try:
            logger.debug(f"Updating cell {cell_ref} with value: {value}")
            cell = self.sheet[cell_ref]
            cell.value = value
            cell.font = self._font
            if wrap_text:
                cell.alignment = Alignment(wrap_text=True)
        except KeyError as e:
            logger.error(f"Invalid cell reference: {cell_ref}")
            raise KeyError(f"Invalid cell reference: {cell_ref}") from e
        except Exception as e:
            logger.error(f"Failed to update cell {cell_ref}", exc_info=True)
            raise

    @log_exceptions(logger)
    def format_date_range(self, start_date: datetime, end_date: datetime) -> str:
        """Format date range for the report using start and end dates"""
        try:
            return f"{start_date.strftime('%m/%d/%y')} - {end_date.strftime('%m/%d/%y')}"
        except Exception as e:
            logger.error("Failed to format date range", exc_info=True)
            # Return a fallback date range if formatting fails
            today = datetime.now()
            month_ago = today - timedelta(days=31)
            return f"{month_ago.strftime('%m/%d/%y')} - {today.strftime('%m/%d/%y')}"

    @log_exceptions(logger)
    def _parse_date(self, date_value: Union[str, datetime, None]) -> Optional[datetime]:
        """
        Parse date from various formats with improved error handling.
        Falls back to current date if parsing fails.
        """
        logger.debug(f"Parsing date value: {date_value} of type {type(date_value)}")

        # Handle None or empty string
        if not date_value:
            return None

        # If already a datetime, return as is
        if isinstance(date_value, datetime):
            return date_value

        if not isinstance(date_value, str):
            logger.warning(f"Unexpected date type: {type(date_value)}")
            return None

        try:
            # Try common formats
            for fmt in [
                '%a, %d %b %Y %H:%M:%S GMT',  # Fri, 22 Nov 2024 00:00:00 GMT
                '%Y-%m-%dT%H:%M:%S.%f',  # ISO format with microseconds
                '%Y-%m-%dT%H:%M:%S',  # ISO format
                '%Y-%m-%d %H:%M:%S',  # Standard datetime
                '%Y-%m-%d'  # Simple date
            ]:
                try:
                    return datetime.strptime(date_value, fmt)
                except ValueError:
                    continue

            logger.warning(f"Could not parse date: {date_value}")
            return None
        except Exception as e:
            logger.error(f"Error parsing date {date_value}: {str(e)}")
            return None

    @log_exceptions(logger)
    def update_metrics(self, metrics: WorkOrderMetrics) -> None:
        """Update the what-if table with work order metrics"""
        if not self.sheet:
            raise ValueError("Workbook not initialized")

        try:
            logger.info("Updating work order metrics and what-if table")

            # Update initial metrics cells
            self.update_cell('B6', 21)  # Days per month
            self.update_cell('M9', metrics.actual_open_work_orders)  # Current actual open work orders

            # Update what-if table with necessary data
            update_what_if_table(
                sheet=self.sheet,
                data={
                    'ActualOpenWorkOrders_Current': metrics.actual_open_work_orders,
                    'CancelledWorkOrder_Current': metrics.cancelled_work_orders,
                    'CompletedWorkOrder_Current': metrics.completed_work_orders,
                    'PendingWorkOrders': metrics.pending_work_orders,
                }
            )

            logger.info("Successfully updated metrics and what-if table")
        except Exception as e:
            logger.error("Failed to update metrics", exc_info=True)
            raise

    @log_exceptions(logger)
    def update_property_data(self, property_data: Dict[str, Any]) -> None:
        """Update cells with property data and handle all calculations"""
        logger.info(f"Updating property data for: {property_data.get('PropertyName', 'Unknown')}")
        try:
            # Extract period dates - they should already be datetime objects
            period_start = property_data.get('period_start_date')
            period_end = property_data.get('period_end_date')

            # Fallback only if dates are somehow missing
            if not period_start or not period_end:
                logger.warning("Missing period dates in property data, using latest_post_date")
                latest_post = property_data.get('latest_post_date')
                if latest_post:
                    period_end = latest_post
                    period_start = latest_post - timedelta(days=30)
                else:
                    logger.warning("No valid dates found, using default date range")
                    period_end = datetime.now()
                    period_start = period_end - timedelta(days=30)

            logger.info(f"Using date range: {period_start} to {period_end}")
            date_range = self.format_date_range(period_start, period_end)

            # Update initial metrics cells
            self.update_cell('M9', property_data.get('ActualOpenWorkOrders_Current', 0))

            # Get break-even value from B24
            break_even_value = self.sheet['B24'].value
            if isinstance(break_even_value, str) and break_even_value.startswith('='):
                from what_if_table import FormulaEvaluator
                evaluator = FormulaEvaluator()
                break_even_value = evaluator.evaluate_formula(break_even_value, self.sheet)
            break_even_value = round(float(break_even_value or 0), 1)
            l12_text = f"Required daily work order output *In addition to Break even\n ({break_even_value} per-workday)*"

            # Update cells with validated data
            updates = {
                'B22': property_data.get('TotalUnitCount', 0),
                'N9': property_data.get('CompletedWorkOrder_Current', 0),
                'O9': property_data.get('PendingWorkOrders', 0),
                'L8': property_data.get('PropertyName', 'Unknown Property'),
                'D4': date_range,
                'M8': date_range,
                'A1': f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                'L12': l12_text
            }

            logger.debug(f"Updating cells with values: {updates}")
            for cell_ref, value in updates.items():
                self.update_cell(cell_ref, value, wrap_text=cell_ref in ['L12'])

            # Update what-if table after cell updates
            update_what_if_table(
                sheet=self.sheet,
                data=property_data
            )

            logger.info("Successfully updated all property data")
        except Exception as e:
            logger.error(f"Failed to update property data: {str(e)}", exc_info=True)
            raise

    @log_exceptions(logger)
    def save(self) -> None:
        """Save workbook with proper error handling and logging"""
        if not self.workbook:
            logger.error("Attempted to save without initialized workbook")
            raise ValueError("No workbook to save")

        try:
            logger.info(f"Saving workbook to: {self.template.output_path}")
            os.makedirs(os.path.dirname(self.template.output_path), exist_ok=True)
            self.workbook.save(self.template.output_path)
            logger.info("Workbook saved successfully")
        except PermissionError:
            logger.error("Permission denied while saving workbook", exc_info=True)
            raise
        except Exception as e:
            logger.error("Failed to save workbook", exc_info=True)
            raise


def generate_multi_property_report(template_name: str, properties: List[Property],
                                   output_dir: Union[str, Path] = None
                                   ) -> List[Path]:
    """
    Generate a single Excel report with multiple sheets for multiple properties.
    Each sheet maintains exact template formatting and formulas.

    Args:
        template_name: Name of the template file
        properties: List of Property objects
        output_dir: Optional output directory path

    Returns:
        List[Path]: List of paths to generated report files
    """
    logger.info(f"Starting multi-property report generation for {len(properties)} properties")

    try:
        # Use provided output directory or create timestamped one
        if output_dir:
            base_output_dir = Path(output_dir)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_output_dir = Path('output') / f'multi_property_report_{timestamp}'

        # Ensure output directory exists
        base_output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Using output directory: {base_output_dir}")

        # Create filename for the consolidated report
        report_filename = f"break_even_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = base_output_dir / report_filename

        # Initialize template
        template = SpreadsheetTemplate(
            template_name=template_name,
            output_path=output_path
        )

        # Load the template once
        template_workbook = openpyxl.load_workbook(template.template_path)
        template_sheet = template_workbook.active

        # Create the service for the consolidated workbook
        service = ExcelGeneratorService(template)
        service.initialize_workbook()

        # Process each property
        for idx, property_data in enumerate(properties):
            try:
                logger.info(f"Processing property {idx + 1}/{len(properties)}: {property_data.property_name}")

                # For the first property, use the existing sheet
                if idx == 0:
                    sheet = service.sheet
                else:
                    # For subsequent properties, create new sheet
                    sheet = service.workbook.create_sheet()

                # Create and clean sheet name
                safe_name = "".join(c if c.isalnum() or c in (' ', '-') else '_' for c in property_data.property_name)
                sheet_name = safe_name[:31]

                # Handle duplicate sheet names
                base_name = sheet_name
                counter = 1
                while sheet_name in (s.title for s in service.workbook.worksheets if s != sheet):
                    suffix = f"_{counter}"
                    sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"
                    counter += 1

                sheet.title = sheet_name

                # Copy template content and formatting
                for row in template_sheet.rows:
                    for cell in row:
                        new_cell = sheet.cell(
                            row=cell.row,
                            column=cell.column
                        )

                        # Copy value and formula
                        if cell.value is not None:
                            new_cell.value = cell.value
                            if cell.data_type == 'f':  # If it's a formula
                                new_cell.data_type = cell.data_type

                        # Copy styling
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.protection = copy(cell.protection)

                # Copy column dimensions
                for key, value in template_sheet.column_dimensions.items():
                    sheet.column_dimensions[key].width = value.width

                # Copy row dimensions
                for key, value in template_sheet.row_dimensions.items():
                    sheet.row_dimensions[key].height = value.height

                # Copy merged cells
                sheet.merged_cells = template_sheet.merged_cells

                # Copy conditional formatting
                sheet.conditional_formatting = template_sheet.conditional_formatting

                # Update service to use current sheet
                service.sheet = sheet

                # Convert Property model to dict format with proper date handling
                property_dict = {
                    'PropertyKey': property_data.property_key,
                    'PropertyName': property_data.property_name,
                    'TotalUnitCount': property_data.total_unit_count,
                    'period_start_date': property_data.period_start_date,
                    'period_end_date': property_data.period_end_date,
                    'latest_post_date': property_data.latest_post_date,
                    'OpenWorkOrder_Current': property_data.metrics.open_work_orders if property_data.metrics else 0,
                    'ActualOpenWorkOrders_Current': property_data.metrics.actual_open_work_orders if property_data.metrics else 0,
                    'CompletedWorkOrder_Current': property_data.metrics.completed_work_orders if property_data.metrics else 0,
                    'CancelledWorkOrder_Current': property_data.metrics.cancelled_work_orders if property_data.metrics else 0,
                    'PendingWorkOrders': property_data.metrics.pending_work_orders if property_data.metrics else 0,
                }

                if property_data.metrics:
                    property_dict.update({
                        'metrics': {
                            'period_start_date': property_data.period_start_date,
                            'period_end_date': property_data.period_end_date,
                            'actual_open_work_orders': property_data.metrics.actual_open_work_orders,
                            'completed_work_orders': property_data.metrics.completed_work_orders,
                            'cancelled_work_orders': property_data.metrics.cancelled_work_orders,
                            'pending_work_orders': property_data.metrics.pending_work_orders,
                        }
                    })

                # Update the sheet with property data
                service.update_property_data(property_dict)

                logger.info(f"Successfully processed {property_data.property_name}")

                time.sleep(1)

            except Exception as e:
                logger.error(f"Failed to process {property_data.property_name}: {str(e)}", exc_info=True)
                continue

        # Save the consolidated workbook
        logger.info(f"Saving consolidated report to {output_path}")
        service.save()

        return [output_path]

    except Exception as e:
        logger.error(f"Failed to generate multi-property report: {str(e)}", exc_info=True)
        if 'base_output_dir' in locals() and base_output_dir.exists():
            try:
                shutil.rmtree(base_output_dir)
                logger.info(f"Cleaned up output directory after failure: {base_output_dir}")
            except Exception as cleanup_error:
                logger.error(f"Failed to clean up output directory: {cleanup_error}")
        raise


def generate_report(template_name: str, output_path: str, property_data: Dict[str, Any]) -> None:
    """Main function to generate Excel report with comprehensive logging"""
    logger.info(f"Starting report generation for property: {property_data.get('PropertyName', 'Unknown')}")

    try:
        # Initialize template with template_name
        template = SpreadsheetTemplate(
            template_name=template_name,
            output_path=Path(output_path)
        )

        # Create service and generate report
        service = ExcelGeneratorService(template)
        service.initialize_workbook()
        service.update_property_data(property_data)
        service.save()

        logger.info("Report generated successfully")
    except Exception as e:
        logger.error("Failed to generate report", exc_info=True)
        raise


if __name__ == "__main__":
    # Example usage with error handling
    template_path = 'break_even_template.xlsx'
    output_path = "output/filled_work_order_report.xlsx"

    # Sample property data
    property_data = {
        "PropertyKey": 1,
        "PropertyName": "Test Property",
        "TotalUnitCount": 100,
        "LatestPostDate": "Wed, 23 Oct 2024 00:00:00 GMT",
        "ActualOpenWorkOrders_Current": 399,
        "CompletedWorkOrder_Current": 200,
        "CancelledWorkOrder_Current": 50,
        "PendingWorkOrders": 58
    }

    try:
        generate_report(template_path, output_path, property_data)
        logger.info("Example report generated successfully!")
    except Exception as e:
        logger.error("Failed to generate example report", exc_info=True)
